# app.py
import streamlit as st
import pandas as pd
import os
import json
import uuid
from datetime import datetime
from openpyxl import load_workbook
from io import BytesIO
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# ----------------------------
# Basic setup
# ----------------------------
st.set_page_config(page_title="📄 Excel→Form + Auto Email + Edit", layout="wide")
st.title("📄 Excel → Auto Form Builder + Auto Email + Edit")

DATA_DIR = "data_store"
FORMS_DIR = os.path.join(DATA_DIR, "forms")
RESP_DIR = os.path.join(DATA_DIR, "responses")
LINKS_FILE = os.path.join(DATA_DIR, "links.json")
os.makedirs(FORMS_DIR, exist_ok=True)
os.makedirs(RESP_DIR, exist_ok=True)
os.makedirs(DATA_DIR, exist_ok=True)

# load or init links mapping
if os.path.exists(LINKS_FILE):
    with open(LINKS_FILE, "r") as f:
        LINKS = json.load(f)
else:
    LINKS = {}
    with open(LINKS_FILE, "w") as f:
        json.dump(LINKS, f, indent=2)

# ----------------------------
# Helpers
# ----------------------------
def read_excel_all_sheets(uploaded_file):
    wb = load_workbook(uploaded_file)
    sheets = {}
    for s in wb.sheetnames:
        sheets[s] = pd.read_excel(uploaded_file, sheet_name=s)
    return sheets

def parse_template_from_df(df_sheets):
    """
    Determine columns and dropdown options.
    Approach:
    - If any header contains "ColName:opt1,opt2", parse that.
    - If there is a sheet called 'Dropdowns' or 'Options', read mappings:
        first column = field name, second column = comma-separated options OR multiple rows per field
    - Otherwise, columns = df.columns of first (or chosen) sheet
    Return: list of fields as dicts: [{"name": "...", "type": "text"|"select", "options": [...]}]
    """
    # prefer sheet named 'Template' or first non-empty sheet
    preferred = None
    for name in ["Template", "Form", "Sheet1"]:
        if name in df_sheets:
            preferred = name
            break
    if not preferred:
        # choose first sheet with tabular columns
        preferred = list(df_sheets.keys())[0]

    df = df_sheets[preferred]
    cols = list(df.columns)

    # parse headers for inline options
    fields = []
    dropdown_map = {}

    # look for a special sheet for dropdowns
    for key in ["Dropdowns", "Options", "Dropdown", "Choices"]:
        if key in df_sheets:
            dd_df = df_sheets[key]
            # assume two-column mapping: field | options (comma-separated) OR multiple rows with same field
            for _, row in dd_df.iterrows():
                if len(row) >= 2 and pd.notna(row.iloc[0]):
                    field_name = str(row.iloc[0]).strip()
                    # options can be in subsequent columns or second column
                    opts = []
                    # collect non-empty values in the row after first column
                    for val in row.iloc[1:]:
                        if pd.notna(val):
                            opts.append(str(val).strip())
                    # if second column is comma-separated
                    if len(opts) == 1 and "," in opts[0]:
                        opts = [o.strip() for o in opts[0].split(",") if o.strip()]
                    if opts:
                        dropdown_map[field_name] = opts
            break

    for c in cols:
        name = str(c).strip()
        # inline format: "Gender:Male,Female"
        if ":" in name:
            parts = name.split(":", 1)
            fld = parts[0].strip()
            opts = [o.strip() for o in parts[1].split(",") if o.strip()]
            fields.append({"name": fld, "type": "select" if opts else "text", "options": opts})
        elif name in dropdown_map:
            fields.append({"name": name, "type": "select", "options": dropdown_map[name]})
        else:
            fields.append({"name": name, "type": "text", "options": []})

    return fields

def save_form_definition(form_id, form_def):
    path = os.path.join(FORMS_DIR, f"{form_id}.json")
    with open(path, "w") as f:
        json.dump(form_def, f, indent=2)

def load_form_definition(form_id):
    path = os.path.join(FORMS_DIR, f"{form_id}.json")
    if os.path.exists(path):
        with open(path, "r") as f:
            return json.load(f)
    return None

def save_response(form_id, resp_id, data):
    path = os.path.join(RESP_DIR, f"{form_id}__{resp_id}.json")
    with open(path, "w") as f:
        json.dump(data, f, indent=2)

def load_responses_for_form(form_id):
    res = []
    for fname in os.listdir(RESP_DIR):
        if fname.startswith(form_id + "__") and fname.endswith(".json"):
            with open(os.path.join(RESP_DIR, fname), "r") as f:
                res.append(json.load(f))
    return res

def generate_token():
    return uuid.uuid4().hex

def persist_links():
    with open(LINKS_FILE, "w") as f:
        json.dump(LINKS, f, indent=2)

def send_email_smtp(sender_email, sender_pass, receiver, subject, body, smtp_server="smtp.gmail.com", smtp_port=587):
    try:
        msg = MIMEMultipart()
        msg["From"] = sender_email
        msg["To"] = receiver
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))

        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, sender_pass)
        server.sendmail(sender_email, receiver, msg.as_string())
        server.quit()
        return True, "Sent"
    except Exception as e:
        return False, str(e)

# ----------------------------
# Sidebar - Email / App config
# ----------------------------
st.sidebar.header("🔧 App & Email Settings")
st.sidebar.markdown("Enter base URL of this app (used to build shareable links). Example: `https://share.streamlit.io/yourname/yourapp`")
APP_BASE_URL = st.sidebar.text_input("App Base URL (required to email links)", value="", help="Enter full base URL without trailing slash")
SENDER_EMAIL = st.sidebar.text_input("Sender Email (for SMTP)")
SENDER_PASS = st.sidebar.text_input("Sender App Password (for SMTP)", type="password")
st.sidebar.markdown("---")
st.sidebar.markdown("Tips:\n- To test without sending real emails, leave sender credentials empty and you'll get links displayed instead of emailed.\n- Dropdowns can be placed as `ColumnName:opt1,opt2` in header or in a separate sheet named 'Dropdowns' or 'Options'.")

# ----------------------------
# Main Tabs
# ----------------------------
tabs = st.tabs(["1. Create Form", "2. Form Link (open/edit)", "3. Dashboard / Responses"])

# ----------------------------
# TAB 1: Create Form
# ----------------------------
with tabs[0]:
    st.header("1️⃣ Create Form from Excel files")

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Upload Members Excel (Name, Email)")
        members_file = st.file_uploader("Upload Members file (.xlsx or .csv)", type=["xlsx", "csv"], key="members_upload")
    with col2:
        st.subheader("Upload Template Excel (columns / dropdowns)")
        template_file = st.file_uploader("Upload Template file (.xlsx or .csv)", type=["xlsx", "csv"], key="template_upload")

    if members_file and template_file:
        # read members
        try:
            if members_file.name.endswith(".csv"):
                df_members = pd.read_csv(members_file)
            else:
                df_members = pd.read_excel(members_file)
            # normalize columns: try to find 'name' and 'email'
            cols_lower = [c.lower() for c in df_members.columns]
            name_col = None
            email_col = None
            for c in df_members.columns:
                lc = c.lower()
                if "name" in lc and name_col is None:
                    name_col = c
                if "email" in lc and email_col is None:
                    email_col = c
            if name_col is None or email_col is None:
                st.error("Couldn't detect 'Name' and 'Email' columns in members file. Make sure columns contain 'name' and 'email' in their header.")
            else:
                st.success(f"Detected members: {len(df_members)} rows. Using Name column `{name_col}` and Email column `{email_col}`")
                # read template
                if template_file.name.endswith(".csv"):
                    df_template = pd.read_csv(template_file)
                    df_sheets = {"Template": df_template}
                else:
                    df_sheets = read_excel_all_sheets(template_file)
                fields = parse_template_from_df(df_sheets)

                st.subheader("Detected Fields")
                st.write("You can edit fields (add / remove / reorder) before creating the form.")
                # store a working copy in session_state
                if "working_fields" not in st.session_state:
                    st.session_state.working_fields = fields.copy()
                    st.session_state.removed_fields = []

                # Show editable list
                for idx, f in enumerate(st.session_state.working_fields):
                    cols_f = st.columns([4, 1, 1, 1])
                    with cols_f[0]:
                        new_name = st.text_input(f"Field {idx+1} name", value=f["name"], key=f"fname_{idx}")
                    with cols_f[1]:
                        field_type = st.selectbox("Type", options=["text", "select"], index=0 if f["type"]=="text" else 1, key=f"ftype_{idx}")
                    with cols_f[2]:
                        opts_val = ""
                        if field_type == "select":
                            opts_val = st.text_input("Options (comma separated)", value=",".join(f.get("options", [])), key=f"fopts_{idx}")
                        else:
                            # empty placeholder
                            st.write("")
                    with cols_f[3]:
                        if st.button("Remove", key=f"remove_{idx}"):
                            st.session_state.removed_fields.append(st.session_state.working_fields.pop(idx))
                            st.experimental_rerun()
                    # update back
                    st.session_state.working_fields[idx]["name"] = new_name
                    st.session_state.working_fields[idx]["type"] = field_type
                    if field_type == "select":
                        st.session_state.working_fields[idx]["options"] = [o.strip() for o in opts_val.split(",") if o.strip()]
                    else:
                        st.session_state.working_fields[idx]["options"] = []

                # option to add new field
                if st.button("Add New Field"):
                    st.session_state.working_fields.append({"name": "New Field", "type": "text", "options": []})
                    st.experimental_rerun()

                # Restore removed
                if st.session_state.removed_fields:
                    st.write("Removed fields:")
                    for ridx, rf in enumerate(st.session_state.removed_fields):
                        if st.button(f"Restore {rf['name']}", key=f"restore_{ridx}"):
                            st.session_state.working_fields.append(rf)
                            st.session_state.removed_fields.pop(ridx)
                            st.experimental_rerun()

                st.markdown("---")
                # Create form button
                if st.button("Create & Save Form"):
                    form_id = uuid.uuid4().hex[:12]
                    form_def = {
                        "form_id": form_id,
                        "created_at": datetime.utcnow().isoformat(),
                        "fields": st.session_state.working_fields,
                        "members_count": len(df_members),
                        "name_col": name_col,
                        "email_col": email_col,
                        "members_preview": df_members.head(10).to_dict(orient="records"),
                    }
                    save_form_definition(form_id, form_def)
                    st.success(f"Form saved with ID: {form_id}")
                    st.write("Next: send links to members or copy links to distribute.")
                    # prepare links mapping for each member
                    LINKS.setdefault(form_id, {})
                    for _, row in df_members.iterrows():
                        recipient = str(row[email_col]).strip()
                        recipient_name = str(row[name_col]).strip()
                        token = generate_token()
                        LINKS[form_id][recipient] = {"name": recipient_name, "token": token, "sent": False}
                    persist_links()
                    st.info("Unique tokens generated for each member and saved. Go to 'Form Link (open/edit)' tab to send links or view them.")

    else:
        st.info("Upload both Members and Template files to create a form.")

# ----------------------------
# TAB 2: Form Link (open/edit)
# ----------------------------
with tabs[1]:
    st.header("2️⃣ Form Link (open or edit responses)")

    # show existing forms
    forms = [f[:-5] for f in os.listdir(FORMS_DIR) if f.endswith(".json")]
    form_choice = st.selectbox("Choose a form to operate on", options=["-- new --"] + forms)
    if form_choice and form_choice != "-- new --":
        fid = form_choice
        form_def = load_form_definition(fid)
        if not form_def:
            st.error("Form definition missing.")
        else:
            st.subheader(f"Form ID: {fid}")
            st.write(f"Created at: {form_def.get('created_at')}")
            st.write("Fields:")
            st.json(form_def.get("fields"))

            # show link generation / sending panel
            st.markdown("### Links for members")
            this_links = LINKS.get(fid, {})
            df_links = pd.DataFrame([{"email": k, "name": v["name"], "token": v["token"], "sent": v.get("sent", False)} for k,v in this_links.items()])
            st.dataframe(df_links)

            base = APP_BASE_URL.strip().rstrip("/")
            if not base:
                st.warning("Set 'App Base URL' in Sidebar to enable emailing direct links. Links will still be shown below for copy/paste.")
            # option to (re)generate tokens for missing
            if st.button("Regenerate tokens for all members"):
                for k in this_links:
                    LINKS[fid][k]["token"] = generate_token()
                    LINKS[fid][k]["sent"] = False
                persist_links()
                st.success("Tokens regenerated.")

            # display / copy links
            st.markdown("#### Member Links (view/edit)")
            for email, info in this_links.items():
                token = info["token"]
                if base:
                    link = f"{base}?form_id={fid}&token={token}"
                else:
                    link = f"?form_id={fid}&token={token}"
                cols_link = st.columns([4,1,1])
                cols_link[0].write(f"**{info['name']}** — {email}")
                cols_link[0].write(link)
                if cols_link[1].button("Copy Link", key=f"copy_{email}"):
                    st.experimental_set_query_params(form_id=fid, token=token)  # assistive, affects user's URL
                    st.success("Set in your browser URL bar (for quick test).")
                if cols_link[2].button("Resend Email", key=f"resend_{email}"):
                    # send email now
                    if not (SENDER_EMAIL and SENDER_PASS and base):
                        st.error("Provide sender email, password and App Base URL in Sidebar to send emails.")
                    else:
                        subject = f"Please fill the form: {fid}"
                        body = f"Hello {info['name']},\n\nPlease fill the form using the link below:\n\n{link}\n\nYou can edit your response using the same link.\n\nThanks."
                        ok, msg = send_email_smtp(SENDER_EMAIL, SENDER_PASS, email, subject, body)
                        if ok:
                            LINKS[fid][email]["sent"] = True
                            persist_links()
                            st.success(f"Email sent to {email}")
                        else:
                            st.error(f"Failed to send: {msg}")

            # bulk send
            if st.button("Send links to ALL members"):
                if not (SENDER_EMAIL and SENDER_PASS and base):
                    st.error("Provide sender email, password and App Base URL in Sidebar to send emails.")
                else:
                    successes = 0
                    failures = []
                    for email, info in this_links.items():
                        link = f"{base}?form_id={fid}&token={info['token']}"
                        subject = f"Please fill the form: {fid}"
                        body = f"Hello {info['name']},\n\nPlease fill the form using the link below:\n\n{link}\n\nYou can edit your response using the same link.\n\nThanks."
                        ok, msg = send_email_smtp(SENDER_EMAIL, SENDER_PASS, email, subject, body)
                        if ok:
                            LINKS[fid][email]["sent"] = True
                            successes += 1
                        else:
                            failures.append({"email": email, "error": msg})
                    persist_links()
                    st.success(f"Emails sent: {successes}. Failures: {len(failures)}")
                    if failures:
                        st.write(failures)

    # Show form when accessed via query params (simulate member visiting link)
    params = st.experimental_get_query_params()
    if "form_id" in params and "token" in params:
        fid = params["form_id"][0]
        token = params["token"][0]
        form_def = load_form_definition(fid)
        if not form_def:
            st.error("Form not found.")
        else:
            # find member by token
            member_email = None
            member_name = None
            for email, info in LINKS.get(fid, {}).items():
                if info.get("token") == token:
                    member_email = email
                    member_name = info.get("name")
                    break
            if not member_email:
                st.error("Invalid token.")
            else:
                st.header(f"Form: {fid} — for {member_name} ({member_email})")
                # check if response already exists for this token (we'll use resp_id = token)
                resp_id = token
                resp_path = os.path.join(RESP_DIR, f"{fid}__{resp_id}.json")
                existing = None
                if os.path.exists(resp_path):
                    with open(resp_path, "r") as f:
                        existing = json.load(f)

                with st.form("member_form"):
                    form_values = {}
                    for fld in form_def["fields"]:
                        fname = fld["name"]
                        if fld["type"] == "select":
                            # if existing value present, set default
                            default = existing.get(fname) if existing else None
                            form_values[fname] = st.selectbox(fname, options=[""] + fld["options"], index=0 if not default else fld["options"].index(default)+1)
                        else:
                            default = existing.get(fname) if existing else ""
                            form_values[fname] = st.text_input(fname, value=default)
                    submitted = st.form_submit_button("Submit / Save")
                    if submitted:
                        payload = {
                            "form_id": fid,
                            "resp_id": resp_id,
                            "email": member_email,
                            "name": member_name,
                            "submitted_at": datetime.utcnow().isoformat(),
                            "answers": form_values
                        }
                        save_response(fid, resp_id, payload)
                        st.success("✅ Response saved.")
                        # send confirmation email to member
                        if SENDER_EMAIL and SENDER_PASS:
                            body = f"Hello {member_name},\n\nWe received your response for form {fid}.\n\nThanks."
                            ok, msg = send_email_smtp(SENDER_EMAIL, SENDER_PASS, member_email, f"Confirmation - Form {fid}", body)
                            if ok:
                                st.info("Confirmation email sent.")
                            else:
                                st.warning(f"Could not send confirmation email: {msg}")
                        else:
                            st.info("No sender credentials provided — confirmation email not sent.")

# ----------------------------
# TAB 3: Dashboard / Responses
# ----------------------------
with tabs[2]:
    st.header("3️⃣ Dashboard - Forms & Responses")
    # list forms
    forms = [f[:-5] for f in os.listdir(FORMS_DIR) if f.endswith(".json")]
    if not forms:
        st.info("No forms created yet.")
    else:
        sel_form = st.selectbox("Select form", options=forms)
        if sel_form:
            form_def = load_form_definition(sel_form)
            st.subheader(f"Form {sel_form}")
            st.write("Fields:")
            st.json(form_def.get("fields"))
            st.write("Members preview:")
            st.write(form_def.get("members_preview", []))

            # list responses
            responses = load_responses_for_form(sel_form)
            if not responses:
                st.info("No responses yet.")
            else:
                # flatten answers into dataframe
                rows = []
                for r in responses:
                    base = {"form_id": r.get("form_id"), "resp_id": r.get("resp_id"), "name": r.get("name"), "email": r.get("email"), "submitted_at": r.get("submitted_at")}
                    answers = r.get("answers", {})
                    base.update(answers)
                    rows.append(base)
                df_resp = pd.DataFrame(rows)
                st.dataframe(df_resp)

                # export CSV
                csv = df_resp.to_csv(index=False).encode("utf-8")
                st.download_button("Download responses as CSV", csv, file_name=f"responses_{sel_form}.csv", mime="text/csv")

                # allow select a response to edit
                resp_ids = df_resp["resp_id"].tolist()
                chosen = st.selectbox("Select response to edit/view", options=resp_ids)
                if chosen:
                    rfile = os.path.join(RESP_DIR, f"{sel_form}__{chosen}.json")
                    if os.path.exists(rfile):
                        with open(rfile, "r") as f:
                            rdata = json.load(f)
                        st.json(rdata)
                        if st.button("Delete this response"):
                            os.remove(rfile)
                            st.success("Response deleted.")
                            st.experimental_rerun()

            # option to delete or reset form
            if st.button("Delete Form (and all responses)"):
                # remove form file
                os.remove(os.path.join(FORMS_DIR, f"{sel_form}.json"))
                # remove responses
                for fname in os.listdir(RESP_DIR):
                    if fname.startswith(sel_form + "__"):
                        os.remove(os.path.join(RESP_DIR, fname))
                # remove links entry
                if sel_form in LINKS:
                    LINKS.pop(sel_form)
                    persist_links()
                st.success("Form and responses deleted.")
                st.experimental_rerun()
