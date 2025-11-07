import streamlit as st
import pandas as pd
import os
import json
import uuid
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import re

# ----------------------------
# Setup
# ----------------------------
st.set_page_config(page_title="📄 Excel → Web Form", layout="centered")
st.title("📄 Auto Form Creator from Excel")

DATA_DIR = "data_store"
os.makedirs(DATA_DIR, exist_ok=True)
META_PATH = os.path.join(DATA_DIR, "meta.json")

# ----------------------------
# Helper Functions
# ----------------------------
def load_meta():
    if os.path.exists(META_PATH):
        with open(META_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_meta(meta):
    with open(META_PATH, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

def detect_dropdowns(excel_file, df_columns):
    excel_file.seek(0)
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active
    dropdowns = {}
    if ws.data_validations:
        for dv in ws.data_validations.dataValidation:
            try:
                if dv.type == "list" and dv.formula1:
                    formula = str(dv.formula1).strip('"')
                    if "," in formula:
                        options = [x.strip() for x in formula.split(",")]
                    else:
                        options = []
                        try:
                            rng = formula.split("!")[-1].replace("$", "")
                            a, b = rng.split(":")
                            col_letters = re.match(r"([A-Za-z]+)", a).group(1)
                            start_row = int(re.match(r"[A-Za-z]+([0-9]+)", a).group(1))
                            end_row = int(re.match(r"[A-Za-z]+([0-9]+)", b).group(1))
                            col_idx = column_index_from_string(col_letters)
                            for r in range(start_row, end_row + 1):
                                v = ws.cell(row=r, column=col_idx).value
                                if v is not None:
                                    options.append(str(v))
                        except Exception:
                            options = []
                    for cell_range in dv.cells:
                        try:
                            cidx = cell_range.min_col - 1
                            if 0 <= cidx < len(df_columns):
                                dropdowns[df_columns[cidx]] = options
                        except Exception:
                            continue
            except Exception:
                continue
    return dropdowns

# ----------------------------
# URL Params
# ----------------------------
params = st.experimental_get_query_params()
mode = params.get("mode", ["admin"])[0]
form_id = params.get("form_id", [None])[0]

meta = load_meta()

# ----------------------------
# FORM VIEW (User fills form)
# ----------------------------

       # ----------------------------
# FORM VIEW (User fills form)
# ----------------------------


        # ----------------------------
# FORM VIEW (User fills form)
# ----------------------------
if mode == "form":
    if not form_id or "forms" not in meta or form_id not in meta["forms"]:
        st.error("Invalid or missing form ID. Please contact the admin.")
    else:
        info = meta["forms"][form_id]
        st.header(f"🧾 {info['form_name']}")

        # ✅ Unique user session for each visitor
        if "session_id" not in st.session_state:
            st.session_state["session_id"] = str(uuid.uuid4())[:8]
        session_id = st.session_state["session_id"]
        st.caption(f"🆔 Your Session ID: {session_id}")

        dropdowns = info.get("dropdowns", {})
        columns = info["columns"]

        # Create form UI dynamically
        values = {}
        for col in columns:
            if col in dropdowns:
                values[col] = st.selectbox(col, dropdowns[col])
            else:
                values[col] = st.text_input(col)

        if st.button("✅ Submit"):
            row = {
                "UserSession": session_id,
                "SubmittedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            row.update(values)

            try:
                original_path = info.get("original_path")

                # ✅ Ensure the Excel file exists, even if missing
                if not original_path or not os.path.exists(original_path):
                    original_path = os.path.join(DATA_DIR, f"original_{form_id}.xlsx")
                    pd.DataFrame(columns=list(row.keys())).to_excel(original_path, index=False)
                    info["original_path"] = original_path
                    meta["forms"][form_id] = info
                    save_meta(meta)

                # ✅ Load existing file
                existing = pd.read_excel(original_path)

                # ✅ Make sure all columns exist
                for col in row.keys():
                    if col not in existing.columns:
                        existing[col] = None

                # ✅ Append safely
                new_row_df = pd.DataFrame([row])
                combined = pd.concat([existing, new_row_df], ignore_index=True)

                # ✅ Write back safely
                combined.to_excel(original_path, index=False)

                st.success("🎉 Your response has been saved successfully!")
                st.balloons()

                # Clear input fields
                for key in values.keys():
                    if key in st.session_state:
                        del st.session_state[key]

            except Exception as e:
                st.error(f"❌ Error saving data: {e}")

               

# ----------------------------
# ADMIN VIEW
# ----------------------------
else:
    st.header("🧑‍💼 Admin Panel")
    st.write("Upload an Excel file — its columns will automatically become form fields.")

    uploaded = st.file_uploader("📂 Upload Excel File", type=["xlsx"])

    if uploaded:
        try:
            df = pd.read_excel(uploaded)
            df.columns = [str(c).strip().replace("_", " ").title() for c in df.columns if pd.notna(c)]
            st.success(f"Detected columns: {len(df.columns)}")
            st.write(df.columns.tolist())

            dropdowns = detect_dropdowns(uploaded, list(df.columns))
            if dropdowns:
                st.info("Detected dropdowns:")
                st.table(pd.DataFrame([{"Field": k, "Options": ", ".join(v)} for k, v in dropdowns.items()]))

            form_name = st.text_input("Form name:", value=f"My Form {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            base_url = st.text_input("Your Streamlit app public URL (example: https://yourapp.streamlit.app)")

            if st.button("🚀 Create Form Link"):
                if not base_url:
                    st.error("Please enter your app URL to generate shareable link.")
                else:
                    form_id = str(uuid.uuid4())[:10]
                    forms = meta.get("forms", {})

                    # ✅ Save uploaded Excel permanently for this form
                    original_path = os.path.join(DATA_DIR, f"original_{form_id}.xlsx")
                    uploaded.seek(0)
                    with open(original_path, "wb") as f:
                        f.write(uploaded.read())

                    forms[form_id] = {
                        "form_name": form_name,
                        "columns": list(df.columns),
                        "dropdowns": dropdowns,
                        "created_at": datetime.now().isoformat(),
                        "original_path": original_path,  # 🔗 Save path to Excel file
                    }

                    meta["forms"] = forms
                    save_meta(meta)

                    link = f"{base_url.rstrip('/')}/?mode=form&form_id={form_id}"
                    st.success("✅ Form created successfully!")
                    st.info("Share this link with others to fill the form:")
                    st.code(link)

        except Exception as e:
            st.error(f"Error processing file: {e}")

        st.markdown("---")
    st.subheader("📊 Existing Forms")

    forms = meta.get("forms", {})

    if forms:
        df_forms = pd.DataFrame([
            {"Form ID": fid, "Form Name": fdata["form_name"], "Created": fdata["created_at"]}
            for fid, fdata in forms.items()
        ])
        st.dataframe(df_forms)

        st.markdown("---")
        st.subheader("📈 Responses Dashboard")

        selected_form = st.selectbox("Select a form to view responses:", ["-- Select --"] + list(forms.keys()))

        if selected_form != "-- Select --":
            fdata = forms[selected_form]
            st.write(f"**Form Name:** {fdata['form_name']}")
            st.write(f"**Created At:** {fdata['created_at']}")

            excel_path = fdata.get("original_path")

            if excel_path and os.path.exists(excel_path):
                try:
                    df_responses = pd.read_excel(excel_path)
                    if not df_responses.empty:
                        st.success(f"✅ {len(df_responses)} responses found")
                        st.dataframe(df_responses, use_container_width=True)

                        # Download button for admin
                        csv = df_responses.to_csv(index=False).encode('utf-8')
                        st.download_button(
                            label="⬇️ Download Responses as CSV",
                            data=csv,
                            file_name=f"{fdata['form_name'].replace(' ', '_')}_responses.csv",
                            mime="text/csv"
                        )
                    else:
                        st.info("No responses submitted yet.")
                except Exception as e:
                    st.error(f"Error reading responses: {e}")
            else:
                st.warning("Excel file not found for this form.")
    else:
        st.info("No forms created yet.")

